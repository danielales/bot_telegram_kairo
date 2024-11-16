import telebot
from telebot import types
import openpyxl

# Token del bot de Telegram
TOKEN = "7527210724:AAFc0kxGluOo_wezme4-rA_i6JYVwAIEYNc"
bot = telebot.TeleBot(TOKEN)

# Función para obtener las zonas únicas desde el archivo Excel
def obtener_zonas():
    archivo = 'BA_quehacer_2.xlsx'
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active
    zonas = set()

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        zona = fila[5]  # Suponiendo que la columna "Zona" es la sexta columna (índice 5)
        if zona:
            zonas.add(zona.strip())  # Agregar zonas eliminando espacios en blanco si los hay

    return sorted(list(zonas))  # Devolver una lista ordenada de zonas únicas

# Función para cargar los eventos en una zona específica
def cargar_eventos_por_zona(zona):
    archivo = 'BA_quehacer_2.xlsx'
    wb = openpyxl.load_workbook(archivo)
    hoja = wb.active
    eventos = []

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        titulo, _, _, _, _, evento_zona = fila  # Extraer título y zona
        if evento_zona and evento_zona.strip().lower() == zona.lower():
            eventos.append(titulo)

    return eventos if eventos else ["No se encontraron eventos en esta zona."]

# Comando /start para mostrar el menú inicial
@bot.message_handler(commands=['start'])
def send_welcome(message):
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    btn1 = types.KeyboardButton("Información")
    btn2 = types.KeyboardButton("Ayuda")
    btn3 = types.KeyboardButton("Contacto")
    btn4 = types.KeyboardButton("Buscar por zona")
    markup.add(btn1, btn2, btn3, btn4)
    bot.send_message(message.chat.id, "¡Bienvenido! Escoge una opción:", reply_markup=markup)

# Muestra los botones de las zonas disponibles
@bot.message_handler(func=lambda message: message.text == "Buscar por zona")
def elegir_zona(message):
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    zonas = obtener_zonas()
    for zona in zonas:
        markup.add(types.KeyboardButton(zona))
    markup.add(types.KeyboardButton("Regresar al menú principal"))
    bot.send_message(message.chat.id, "Selecciona una zona:", reply_markup=markup)

# Responde al seleccionar una zona
@bot.message_handler(func=lambda message: message.text in obtener_zonas())
def mostrar_eventos_zona(message):
    zona = message.text
    eventos = cargar_eventos_por_zona(zona)
    respuesta = f"Eventos en {zona}:\n\n" + "\n".join([f"- {evento}" for evento in eventos])
    bot.send_message(message.chat.id, respuesta)

# Otras opciones del menú principal
@bot.message_handler(func=lambda message: True)
def menu_response(message):
    if message.text == "Información":
        bot.reply_to(message, "Aquí tienes información útil sobre el bot.")
    elif message.text == "Ayuda":
        bot.reply_to(message, "Estoy aquí para ayudarte. ¿En qué necesitas asistencia?")
    elif message.text == "Contacto":
        bot.reply_to(message, "Puedes contactarnos en: contacto@miempresa.com")
    elif message.text == "Regresar al menú principal":
        send_welcome(message)
    else:
        bot.reply_to(message, "Por favor, selecciona una opción válida.")

# Iniciar el bot
bot.polling()